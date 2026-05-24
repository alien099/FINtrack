import io
from datetime import datetime, timedelta, date
from django.conf import settings
from django.db import models
from django.db.models import Sum, F, Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from rest_framework import viewsets, generics
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated, AllowAny

from .models import Category, Transaction, Budget, Account
from django.contrib.auth.models import User
from .serializers import (
    CategorySerializer,
    TransactionSerializer,
    BudgetSerializer,
    RegisterSerializer,
    AccountSerializer
)

class NoPagination(PageNumberPagination):
    page_size = None


class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        return Response({"username": request.user.username})


# В RegisterView после создания пользователя:
class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = [AllowAny]

    def perform_create(self, serializer):
        user = serializer.save()

        # Счет по умолчанию
        Account.objects.create(
            user=user,
            name='Наличные',
            type='cash',
            is_default=True
        )

        # Категория "Перевод" для расходов
        Category.objects.get_or_create(
            name='Перевод',
            type='expense',
            user=user
        )

        # Категория "Перевод" для доходов
        Category.objects.get_or_create(
            name='Перевод',
            type='income',
            user=user
        )

class AccountViewSet(viewsets.ModelViewSet):
    serializer_class = AccountSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Account.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def perform_destroy(self, instance):
        active_count = Account.objects.filter(
            user=self.request.user,
            is_active=True
        ).count()

        if active_count <= 1:
            return Response(
                {"error": "Нельзя удалить единственный счет. Должен быть хотя бы один активный счет."},
                status=400
            )

        another_account = Account.objects.filter(
            user=self.request.user,
            is_active=True
        ).exclude(id=instance.id).first()

        if another_account:
            Transaction.objects.filter(account=instance).update(account=another_account)

        instance.delete()

class TransferView(APIView):
    """Перевод между счетами"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        from_account_id = request.data.get("from_account")
        to_account_id = request.data.get("to_account")
        amount = request.data.get("amount")

        if not from_account_id or not to_account_id:
            return Response({"error": "Выберите оба счета"}, status=400)
        if from_account_id == to_account_id:
            return Response({"error": "Счета должны быть разными"}, status=400)
        if not amount or float(amount) <= 0:
            return Response({"error": "Введите положительную сумму"}, status=400)

        try:
            from_account = Account.objects.get(id=from_account_id, user=request.user, is_active=True)
            to_account = Account.objects.get(id=to_account_id, user=request.user, is_active=True)
        except Account.DoesNotExist:
            return Response({"error": "Счет не найден"}, status=404)

        amount = float(amount)

        # Ищем дефолтные категории перевода
        exp_cat = Category.objects.filter(
            name='Перевод со счета',
            type='expense',
            is_default=True
        ).first()

        inc_cat = Category.objects.filter(
            name='Перевод на счет',
            type='income',
            is_default=True
        ).first()

        if not exp_cat or not inc_cat:
            return Response({"error": "Категории перевода не найдены"}, status=500)

        # Расход со счета
        Transaction.objects.create(
            user=request.user,
            account=from_account,
            amount=amount,
            currency='RUB',
            category=exp_cat,
            date=datetime.now().date(),
            description=f"Перевод на «{to_account.name}»"
        )

        # Доход на счет
        Transaction.objects.create(
            user=request.user,
            account=to_account,
            amount=amount,
            currency='RUB',
            date=datetime.now().date(),
            category=inc_cat,
            description=f"Перевод от «{from_account.name}»"
        )

        return Response({"status": "ok"})
    
class CategoryViewSet(viewsets.ModelViewSet):
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Category.objects.filter(
            models.Q(user=self.request.user) |
            models.Q(is_default=True)
        ).order_by("name")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    def destroy(self, request, *args, **kwargs):
        category = self.get_object()

        if category.is_default:
            return Response(
                {"error": "Нельзя удалить стандартную категорию"},
                status=400
            )

        fallback_id = settings.FALLBACK_IDS.get(category.type)
        if not fallback_id:
            return Response(
                {"error": f"Нет fallback для типа {category.type}"},
                status=500
            )

        try:
            fallback = Category.objects.get(id=fallback_id)
        except Category.DoesNotExist:
            return Response(
                {"error": "Fallback категория не найдена в базе"},
                status=500
            )

        Transaction.objects.filter(
            category=category,
            user=request.user
        ).update(category=fallback)

        Budget.objects.filter(
            category=category,
            user=request.user
        ).update(category=fallback)

        category.delete()
        return Response(status=204)


class TransactionViewSet(viewsets.ModelViewSet):
    serializer_class = TransactionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        qs = Transaction.objects.filter(user=self.request.user)

        start = self.request.query_params.get("date_from")
        end = self.request.query_params.get("date_to")
        ttype = self.request.query_params.get("type")
        categories = self.request.query_params.get("categories")
        account = self.request.query_params.get("account")

        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)
        if ttype:
            qs = qs.filter(category__type=ttype)
        if categories:
            category_ids = [int(x) for x in categories.split(",") if x.isdigit()]
            qs = qs.filter(category_id__in=category_ids)
        if account:
            qs = qs.filter(account_id=account)

        return qs.order_by("-date", "-id")

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class ExportTransactionsView(APIView):
    """Экспорт всех транзакций в Excel"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        transactions = Transaction.objects.filter(user=request.user).select_related('category', 'account')

        wb = Workbook()
        ws = wb.active
        ws.title = "Транзакции"

        headers = ["Дата", "Счет", "Сумма", "Тип", "Категория", "Описание"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0A315C", end_color="0A315C", fill_type="solid")
        header_align = Alignment(horizontal="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        for t in transactions:
            ws.append([
                t.date.strftime("%d.%m.%Y"),
                t.account.name if t.account else "",
                float(t.amount),
                "доход" if t.category and t.category.type == "income" else "расход",
                t.category.name if t.category else "",
                t.description or ""
            ])

        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 30

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="transactions.xlsx"'
        wb.save(response)
        return response


class DownloadTemplateView(APIView):
    """Скачать шаблон для импорта транзакций"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Шаблон импорта"

        headers = ["Дата (ДД.ММ.ГГГГ)", "Счет", "Сумма", "Тип (доход/расход)", "Категория", "Описание"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0A315C", end_color="0A315C", fill_type="solid")
        header_align = Alignment(horizontal="center")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        examples = [
            ["01.01.2026", "Наличные", "5000", "доход", "Зарплата", "Аванс за январь"],
            ["05.01.2026", "Наличные", "1200", "расход", "Продукты", "Пятерочка"],
            ["10.01.2026", "Наличные", "350", "расход", "Транспорт", "Метро"],
        ]
        for row in examples:
            ws.append(row)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 30

        ws.freeze_panes = 'A2'

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="template_import.xlsx"'
        wb.save(response)
        return response


class ImportTransactionsView(APIView):
    """Импорт транзакций из Excel со строгой проверкой"""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({"error": "Файл не найден"}, status=400)

        try:
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active

            imported = 0
            errors = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    continue

                try:
                    # === ОБРАБОТКА ДАТЫ ===
                    raw_date = row[0]
                    if raw_date is None or (isinstance(raw_date, str) and not raw_date.strip()):
                        errors.append(f"Строка {row_idx}: не указана дата")
                        continue

                    if isinstance(raw_date, datetime):
                        date_val = raw_date.date()
                    elif isinstance(raw_date, date):
                        date_val = raw_date
                    elif isinstance(raw_date, (int, float)):
                        try:
                            date_val = (datetime(1899, 12, 30) + timedelta(days=int(raw_date))).date()
                        except:
                            errors.append(f"Строка {row_idx}: неверный формат даты «{raw_date}»")
                            continue
                    else:
                        date_str = str(raw_date).strip()
                        parsed = False
                        for fmt in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
                            try:
                                date_val = datetime.strptime(date_str, fmt).date()
                                parsed = True
                                break
                            except ValueError:
                                continue
                        if not parsed:
                            errors.append(f"Строка {row_idx}: неверный формат даты «{date_str}», нужно ДД.ММ.ГГГГ")
                            continue

                    # === ОСТАЛЬНЫЕ ПОЛЯ ===
                    account_name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                    amount_str = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                    trans_type = str(row[3]).strip().lower() if len(row) > 3 and row[3] else ""
                    category_name = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                    description = str(row[5]).strip() if len(row) > 5 and row[5] else ""

                    # Проверка суммы
                    if not amount_str:
                        errors.append(f"Строка {row_idx}: не указана сумма")
                        continue
                    try:
                        amount = float(amount_str)
                    except ValueError:
                        errors.append(f"Строка {row_idx}: неверная сумма «{amount_str}»")
                        continue
                    if amount <= 0:
                        errors.append(f"Строка {row_idx}: сумма должна быть положительной")
                        continue

                    # Проверка типа
                    if trans_type not in ['доход', 'расход', 'income', 'expense']:
                        errors.append(f"Строка {row_idx}: неверный тип «{trans_type}», укажите «доход» или «расход»")
                        continue
                    category_type = 'income' if trans_type in ['доход', 'income'] else 'expense'

                    # Проверка категории
                    if not category_name:
                        errors.append(f"Строка {row_idx}: не указана категория")
                        continue

                    category = Category.objects.filter(
                        name__iexact=category_name,
                        type=category_type
                    ).filter(
                        models.Q(user=request.user) | models.Q(is_default=True)
                    ).first()

                    if not category:
                        errors.append(
                            f"Строка {row_idx}: категория «{category_name}» с типом «{trans_type}» не найдена. "
                            f"Сначала создайте её в приложении"
                        )
                        continue

                    # Проверка счета
                    if account_name:
                        account = Account.objects.filter(
                            user=request.user,
                            name__iexact=account_name,
                            is_active=True
                        ).first()

                        if not account:
                            errors.append(
                                f"Строка {row_idx}: счет «{account_name}» не найден. "
                                f"Сначала создайте его в приложении"
                            )
                            continue
                    else:
                        account = Account.objects.filter(
                            user=request.user,
                            is_active=True
                        ).first()

                    # Создание транзакции
                    Transaction.objects.create(
                        user=request.user,
                        amount=amount,
                        currency='RUB',
                        date=date_val,
                        category=category,
                        account=account,
                        description=description
                    )
                    imported += 1

                except Exception as e:
                    errors.append(f"Строка {row_idx}: ошибка обработки — {str(e)}")

            return Response({
                "imported": imported,
                "total_errors": len(errors),
                "errors": errors[:20]
            })

        except Exception as e:
            return Response({"error": f"Ошибка чтения файла: {str(e)}"}, status=400)

class BalanceView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Transaction.objects.filter(user=request.user)

        income = qs.filter(category__type='income').aggregate(
            total=Sum('amount')
        )['total'] or 0

        expense = qs.filter(category__type='expense').aggregate(
            total=Sum('amount')
        )['total'] or 0

        return Response({
            "balance": income - expense,
            "income": income,
            "expense": expense
        })


class BudgetViewSet(viewsets.ModelViewSet):
    serializer_class = BudgetSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return Budget.objects.filter(user=self.request.user)

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class AnalyticsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        qs = Transaction.objects.filter(user=request.user)
        start = request.query_params.get("date_from")
        end = request.query_params.get("date_to")
        account = request.query_params.get("account")
        group_by = request.query_params.get("group_by", "month")

        if start: qs = qs.filter(date__gte=start)
        if end: qs = qs.filter(date__lte=end)
        if account: qs = qs.filter(account_id=account)

        if group_by == "day":
            monthly = qs.values("date").annotate(
                income=Sum("amount", filter=Q(category__type="income")),
                expense=Sum("amount", filter=Q(category__type="expense"))
            ).order_by("date")
            monthly = [{"month": str(m["date"]), "income": m["income"] or 0, "expense": m["expense"] or 0} for m in monthly]
        else:
            monthly = qs.annotate(month=TruncMonth("date")).values("month").annotate(
                income=Sum("amount", filter=Q(category__type="income")),
                expense=Sum("amount", filter=Q(category__type="expense"))
            ).order_by("month")
            monthly = [{"month": m["month"].strftime("%Y-%m"), "income": m["income"] or 0, "expense": m["expense"] or 0} for m in monthly]

        expense_structure = qs.filter(category__type="expense").values(name=F("category__name")).annotate(total=Sum("amount")).order_by("-total")
        income_structure = qs.filter(category__type="income").values(name=F("category__name")).annotate(total=Sum("amount")).order_by("-total")

        return Response({
            "monthly": list(monthly),
            "expense_structure": list(expense_structure),
            "income_structure": list(income_structure),
        })